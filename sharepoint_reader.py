import os
import io
import re
import logging
import tempfile
import requests

log = logging.getLogger(__name__)

TENANT_ID  = os.environ.get("GRAPH_TENANT_ID",  "d0f025f6-6cda-470f-be0c-2f50c564a639")
CLIENT_ID  = os.environ.get("GRAPH_CLIENT_ID",  "47e5dc4d-826f-4a25-a00f-988ed032661a")
THUMBPRINT = os.environ.get("GRAPH_THUMB",      "E75578C7AA2C5CE54C326D9DD9F96D0E0DFC9698")

# PEM key — two modes:
#   Local:   GRAPH_PEM_KEY = file path e.g. C:/Users/.../PinnacleLeadPoller_key.pem
#   Railway: GRAPH_PEM_CONTENT = full PEM content as env var (multiline)
PEM_KEY_PATH    = os.environ.get("GRAPH_PEM_KEY", "")
PEM_KEY_CONTENT = os.environ.get("GRAPH_PEM_CONTENT", "")

SHAREPOINT_USER = "kripa@bharathimeraki.com"
FILE_NAME       = "E4 - Master.xlsx"
GRAPH_BASE      = "https://graph.microsoft.com/v1.0"


def _get_pem_bytes() -> bytes:
    # Railway mode: content is directly in env var
    if PEM_KEY_CONTENT:
        return PEM_KEY_CONTENT.encode("utf-8")
    # Local mode: read from file path
    if PEM_KEY_PATH and os.path.exists(PEM_KEY_PATH):
        with open(PEM_KEY_PATH, "rb") as f:
            return f.read()
    # Auto-detect local default path
    default = "C:/Users/bharathimeraki/Downloads/PinnacleLeadPoller_key.pem"
    if os.path.exists(default):
        with open(default, "rb") as f:
            return f.read()
    raise RuntimeError(
        "PEM key not found. Set GRAPH_PEM_CONTENT (Railway) "
        "or GRAPH_PEM_KEY (local file path)."
    )


def _get_access_token():
    import msal
    pem_data = _get_pem_bytes()
    app = msal.ConfidentialClientApplication(
        client_id=CLIENT_ID,
        authority="https://login.microsoftonline.com/" + TENANT_ID,
        client_credential={
            "private_key": pem_data,
            "thumbprint":  THUMBPRINT,
        },
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" in result:
        log.info("Graph API token acquired")
        return result["access_token"]
    raise RuntimeError("Token failed: " + str(result.get("error_description", "")))


def _get(path, token):
    r = requests.get(GRAPH_BASE + path,
                     headers={"Authorization": "Bearer " + token}, timeout=30)
    r.raise_for_status()
    return r.json()


def _get_bytes(path, token):
    r = requests.get(GRAPH_BASE + path,
                     headers={"Authorization": "Bearer " + token}, timeout=60)
    r.raise_for_status()
    return r.content


def _normalise_phone(raw):
    if not raw:
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    if digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]
    return digits


def _extract_crm_id(value):
    if not value:
        return ""
    m = re.search(r'\(#(\d+)\)', str(value))
    if m:
        return m.group(1)
    c = str(value).strip()
    return c if c.isdigit() else ""


def _parse_excel(xlsx_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    header_row_idx = None
    headers = {}

    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        for cell in row:
            if cell and "lead" in str(cell).lower() and "id" in str(cell).lower():
                header_row_idx = row_idx
                break
        if header_row_idx:
            for col_idx, cell in enumerate(
                next(ws.iter_rows(min_row=header_row_idx,
                                  max_row=header_row_idx, values_only=True))
            ):
                if cell:
                    headers[str(cell).strip().lower()] = col_idx
            break

    if header_row_idx is None:
        log.error("Header row not found")
        return {}

    log.info("Headers: " + str(list(headers.keys())[:10]))

    def find_col(*names):
        for n in names:
            if n.lower() in headers:
                return headers[n.lower()]
        return None

    col_id    = find_col("lead's id", "lead id", "id", "crm id")
    col_fname = find_col("first name", "firstname")
    col_lname = find_col("last name", "lastname")
    col_name  = find_col("name", "lead name")
    col_phone = find_col("phone", "mobile", "phone number", "mobile number",
                         "contact number", "phone no")
    col_email = find_col("email", "email id", "email address")

    log.info("col_id=%s col_phone=%s col_email=%s" % (col_id, col_phone, col_email))

    if col_phone is None:
        log.warning("Phone column not found — check sheet headers")

    contacts = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue
        crm_id = ""
        if col_id is not None and col_id < len(row):
            crm_id = _extract_crm_id(str(row[col_id] or ""))
        if not crm_id:
            continue

        name = ""
        if col_name is not None and col_name < len(row):
            name = str(row[col_name] or "").strip()
        elif col_fname is not None:
            fname = str(row[col_fname] or "").strip() if col_fname < len(row) else ""
            lname = str(row[col_lname] or "").strip() if col_lname and col_lname < len(row) else ""
            name  = (fname + " " + lname).strip()

        phone = ""
        if col_phone is not None and col_phone < len(row):
            phone = _normalise_phone(str(row[col_phone] or ""))

        email = ""
        if col_email is not None and col_email < len(row):
            email = str(row[col_email] or "").strip()

        contacts[crm_id] = {"phone": phone, "name": name, "email": email}

    log.info("Parsed %d rows. With phone: %d" % (
        len(contacts), sum(1 for c in contacts.values() if c["phone"])))
    return contacts


def fetch_contacts_from_sharepoint():
    token    = _get_access_token()
    drive    = _get("/users/" + SHAREPOINT_USER + "/drive", token)
    drive_id = drive["id"]
    log.info("Drive: " + drive_id)

    search   = _get("/drives/" + drive_id + "/root/search(q='" + FILE_NAME + "')", token)
    items    = search.get("value", [])
    if not items:
        raise RuntimeError("File not found: " + FILE_NAME)

    file_id  = items[0]["id"]
    log.info("File: " + items[0]["name"] + " id=" + file_id)

    xlsx     = _get_bytes("/drives/" + drive_id + "/items/" + file_id + "/content", token)
    log.info("Downloaded %d bytes" % len(xlsx))

    return _parse_excel(xlsx)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s")
    contacts = fetch_contacts_from_sharepoint()
    print("\nTotal: %d" % len(contacts))
    print("With phone: %d" % sum(1 for c in contacts.values() if c["phone"]))
    print("\nSample (first 5):")
    for crm_id, d in list(contacts.items())[:5]:
        print("  #%s: %s | phone=%s | email=%s" % (crm_id, d["name"], d["phone"], d["email"]))
